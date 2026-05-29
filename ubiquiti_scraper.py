#!/usr/bin/env python3
"""
ubiquiti_scraper.py  v3.0
=========================
Source unique : eu.store.ui.com
Balaye 17 pages catégorie, visite uniquement les fiches avec variantes
ou les pages collection. Chaque variante = une ligne distincte.
Pas de doublon par URL.

Installation
    pip install playwright openpyxl
    playwright install chromium

Usage
    python ubiquiti_scraper.py
    python ubiquiti_scraper.py --output D:\\SRP\\produits.xlsx
    python ubiquiti_scraper.py --no-variants   # ignore les variantes (plus rapide)
    python ubiquiti_scraper.py --debug         # sauve snapshots HTML par catégorie
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import time
from collections import defaultdict
from datetime import date as _date
from pathlib import Path
from typing import Dict, List, Tuple

# ─────────────────────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────────────────────

SCRIPT_DIR     = Path(__file__).parent
DEFAULT_OUTPUT = str(SCRIPT_DIR / "ubiquiti_products.xlsx")

STORE_BASE = "https://eu.store.ui.com"

STORE_CATEGORIES: List[Tuple[str, str]] = [
    ("Cloud Gateways",      f"{STORE_BASE}/eu/en/category/all-cloud-gateways"),
    ("Switching",           f"{STORE_BASE}/eu/en/category/all-switching"),
    ("WiFi",                f"{STORE_BASE}/eu/en/category/all-wifi"),
    ("Cameras & NVR",       f"{STORE_BASE}/eu/en/category/all-cameras-nvrs"),
    ("Door Access",         f"{STORE_BASE}/eu/en/category/all-door-access"),
    ("Integrations",        f"{STORE_BASE}/eu/en/category/all-integrations"),
    ("Advanced Hosting",    f"{STORE_BASE}/eu/en/category/all-advanced-hosting"),
    ("Acc. Cables & DACs",  f"{STORE_BASE}/eu/en/category/accessories-cables-dacs"),
    ("Acc. Modules Fiber",  f"{STORE_BASE}/eu/en/category/accessories-modules-fiber"),
    ("Acc. SFP",            f"{STORE_BASE}/eu/en/category/accessories-sfp-liberation-day"),
    ("Acc. Storage",        f"{STORE_BASE}/eu/en/category/accessories-storage"),
    ("Acc. Rack Mount",     f"{STORE_BASE}/eu/en/category/accessories-rack-mount"),
    ("Acc. PoE & Power",    f"{STORE_BASE}/eu/en/category/accessories-poe-power"),
    ("Acc. Access Point",   f"{STORE_BASE}/eu/en/category/accessories-access-point"),
    ("Acc. Camera",         f"{STORE_BASE}/eu/en/category/accessories-camera"),
    ("Acc. Door Access",    f"{STORE_BASE}/eu/en/category/accessories-door-access"),
    ("Acc. Installations",  f"{STORE_BASE}/eu/en/category/accessories-installations"),
]

BROWSER_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/124.0.0.0 Safari/537.36"
)

CSV_FIELDS = [
    "product_code",
    "product_name",
    "variant_label",
    "category",
    "price_eur",
    "description",
    "available",
    "sold_out",
    "product_url",
]


# ─────────────────────────────────────────────────────────────────────────────
# Utilitaires
# ─────────────────────────────────────────────────────────────────────────────

def _norm_url(url: str) -> str:
    """Supprime le query string et le slash final pour la déduplication."""
    return url.split("?")[0].rstrip("/")


def _parse_price(text: str) -> str:
    """
    Extrait la valeur numérique d'un texte prix EU.

    Exemples :
        "2.000,00 €"          → "2000.00"
        "264,00 €"            → "264.00"
        "1,80 €"              → "1.80"
        "From 264,00 €"       → "264.00"
        "2.400,00 € VAT incl."→ "2400.00"
        "264.00"              → "264.00"   (JSON-LD, déjà numérique)
    """
    if not text:
        return ""
    t = text.strip()
    # Remplace l'espace insécable &nbsp; (U+00A0) par un espace normal
    t = t.replace("\xa0", " ")
    # Supprime préfixe "From" / "À partir de"
    t = re.sub(r"^(from|à partir de)\s*", "", t, flags=re.IGNORECASE)
    # Supprime unités / labels de TVA
    t = re.sub(r"(VAT\s*incl\.?|TTC|HT|€|\$|USD|EUR)", "", t, flags=re.IGNORECASE)
    t = t.strip()

    # Format EU avec séparateur milliers : "2.000,00"
    if re.search(r"\d{1,3}\.\d{3},\d{2}", t):
        t = t.replace(".", "").replace(",", ".")
    # Format EU sans milliers : "264,00"
    elif re.search(r"^\d+,\d{1,2}$", t):
        t = t.replace(",", ".")
    # Format EU avec espaces milliers : "2 000,00"
    elif re.search(r"\d{1,3}\s\d{3},\d{2}", t):
        t = re.sub(r"\s", "", t).replace(",", ".")
    # Format US avec milliers : "2,000.00"
    elif re.search(r"\d{1,3},\d{3}\.\d{2}", t):
        t = t.replace(",", "")

    # Garde uniquement chiffres et point décimal
    t = re.sub(r"[^\d.]", "", t)
    # Évite un point résiduel isolé
    if t in (".", ""):
        return ""
    return t


def _strip_html(html: str) -> str:
    text = re.sub(r"<[^>]+>", " ", html or "")
    text = re.sub(r"\s+", " ", text).strip()
    for ent, rep in [("&amp;", "&"), ("&lt;", "<"), ("&gt;", ">"),
                     ("&nbsp;", " "), ("&#39;", "'"), ("&quot;", '"')]:
        text = text.replace(ent, rep)
    return text


# ─────────────────────────────────────────────────────────────────────────────
# JavaScript injecté – page catégorie / collection
#
# Sélecteurs issus de l'HTML réel du store (mai 2025) :
#   Carte produit  : a.sc-13lrnwl-14
#   Nom commercial : .sc-13lrnwl-4
#   Code / SKU     : .sc-13lrnwl-5
#   Description    : .sc-13lrnwl-7
#   Prix HT        : .sc-13lrnwl-19 (câbles) | .sc-13lrnwl-23 (HDDs/storage)
#                    fallback stable cross-produit : .sc-bw6p3d-16
#   Prix TTC       : .sc-13lrnwl-21 (câbles) | .sc-13lrnwl-25 (HDDs/storage)
#                    fallback stable cross-produit : .sc-bw6p3d-17
#   Badge Sold Out : .sc-1cuaoi0-0    (contient le texte "Sold Out")
#   Bouton action  : button[label]    valeurs: "Add to Cart" | "Select" | "Sold Out"
#
# Layout fiche produit détaillée (sc-pr5ovy-* component) :
#   Nom        : .sc-pr5ovy-5
#   SKU        : .sc-pr5ovy-7
#   Prix       : .sc-pr5ovy-14
#   Description: .sc-14cjdti-2 p
#   Variantes type (Enterprise/Basic) : .sc-pr5ovy-19 .sc-14eyr3g-0.KXSod button[title]
#   Variantes taille (8TB/16TB/24TB)  : .sc-14eyr3g-0.kmiDLq button[title]
# ─────────────────────────────────────────────────────────────────────────────

_JS_CATEGORY_CARDS = r"""() => {
    const results = [];
    const seen = new Set();

    // ── Extraction commune ────────────────────────────────────────────────
    // anchor : l'élément <a> (contient nom, SKU, description)
    // scope  : conteneur racine de la carte (contient aussi prix + boutons
    //          pour le Layout 2 où ils sont hors de l'ancre)
    // Retourne le premier texte non-vide parmi une liste de sélecteurs CSS.
    // Contrairement à querySelector(A)||querySelector(B), cette fonction vérifie
    // le contenu textuel — un nœud DOM vide est ignoré même s'il existe.
    function txt(scope, ...sels) {
        for (const s of sels) {
            const t = scope.querySelector(s)?.innerText?.trim();
            if (t) return t;
        }
        return '';
    }

    function processCard(anchor, scope) {
        const href = anchor.href ? anchor.href.split('?')[0].replace(/\/$/, '') : '';
        if (!href || seen.has(href)) return;
        seen.add(href);

        const name  = txt(anchor, '.sc-13lrnwl-4');
        const sku   = txt(anchor, '.sc-13lrnwl-5');

        // Description : sc-13lrnwl-7 (layout 1) ou sc-bw6p3d-13 (layout 2)
        const desc  = txt(anchor, '.sc-13lrnwl-7', '.sc-bw6p3d-13');

        // Prix HT — priorité :
        //   sc-13lrnwl-23 : produits principaux (WiFi, Switch, GW…) ET HDDs
        //   sc-13lrnwl-19 : câbles/accessoires (sc-bw6p3d layout 2)
        //   sc-bw6p3d-16  : fallback stable Layout 2 (même élément que 19 ou 23)
        // ⚠ Ne PAS utiliser || sur les nœuds DOM : un div vide est truthy en JS.
        //   La fonction txt() vérifie le contenu textuel avant de valider.
        const priceHT  = txt(scope, '.sc-13lrnwl-23', '.sc-13lrnwl-19', '.sc-bw6p3d-16');
        // Prix TTC : même logique
        const priceTTC = txt(scope, '.sc-13lrnwl-25', '.sc-13lrnwl-21', '.sc-bw6p3d-17');

        // Badge "Sold Out"
        const soldEl  = scope.querySelector('.sc-1cuaoi0-0');
        const soldOut = soldEl ? soldEl.innerText.toLowerCase().includes('sold') : false;

        // Bouton action
        const btn      = scope.querySelector('button[label]');
        const btnLabel = btn ? btn.getAttribute('label') : '';

        // Boutons variante dans toute la carte (longueur, couleur, type…)
        let hasVariants = false;
        for (const ctr of scope.querySelectorAll('.sc-14eyr3g-0')) {
            if (ctr.querySelectorAll('button[title]').length > 1) {
                hasVariants = true;
                break;
            }
        }

        const needsVisit = (btnLabel === 'Select')
                        || href.includes('/collections/')
                        || hasVariants;
        const available  = !(soldOut || btnLabel === 'Sold Out');

        if (name || sku) {
            results.push({ href, name, sku, desc, priceHT, priceTTC,
                           soldOut, available, needsVisit, btnLabel });
        }
    }

    // ── Layout 1 : a.sc-13lrnwl-14 (produits principaux) ─────────────────
    // Prix et bouton sont à l'intérieur de l'ancre → scope = ancre
    for (const anchor of document.querySelectorAll('a.sc-13lrnwl-14')) {
        processCard(anchor, anchor);
    }

    // ── Layout 2 : a.sc-bw6p3d-7 (accessoires / câbles) ──────────────────
    // Prix et boutons sont hors de l'ancre → remonter au conteneur .sc-bw6p3d-0
    for (const anchor of document.querySelectorAll('a.sc-bw6p3d-7')) {
        const root = anchor.closest('.sc-bw6p3d-0')
                  || anchor.parentElement?.parentElement
                  || anchor;
        processCard(anchor, root);
    }

    return results;
}"""


# ─────────────────────────────────────────────────────────────────────────────
# JavaScript injecté – fiche produit individuelle
#
# Priorité : JSON-LD (structured data) > sélecteurs DOM
# Variantes : boutons dans .sc-14eyr3g-0 > button[title]
# ─────────────────────────────────────────────────────────────────────────────

_JS_PRODUCT_DATA = r"""() => {
    // 1. JSON-LD structured data (le plus fiable, indépendant du CSS)
    let ldSku = '', ldName = '', ldDesc = '', ldPrice = '';
    for (const s of document.querySelectorAll('script[type="application/ld+json"]')) {
        try {
            const d = JSON.parse(s.textContent);
            const candidates = d['@type'] === 'Product'
                ? [d]
                : (Array.isArray(d['@graph'])
                    ? d['@graph'].filter(x => x && x['@type'] === 'Product')
                    : []);
            for (const prod of candidates) {
                ldName  = ldName  || (prod.name        || '');
                ldDesc  = ldDesc  || (prod.description || '');
                ldSku   = ldSku   || (prod.sku || prod.model || '');
                const offer = Array.isArray(prod.offers)
                    ? prod.offers[0]
                    : (prod.offers || {});
                ldPrice = ldPrice || String(offer.price || '');
            }
        } catch(e) {}
    }

    // 2. Sélecteurs DOM — essaie dans l'ordre :
    //    a) composant carte listing (sc-13lrnwl-*) présent si la fiche réutilise ce composant
    //    b) composant fiche produit détaillée (sc-pr5ovy-*)
    //    c) fallback HTML sémantique (h1 pour le nom)
    function first(selectors) {
        for (const s of selectors) {
            const el = document.querySelector(s);
            if (el) {
                const t = el.innerText ? el.innerText.trim() : el.textContent.trim();
                if (t) return t;
            }
        }
        return '';
    }

    const domSku   = first(['.sc-13lrnwl-5',  '.sc-pr5ovy-7']);
    const domName  = first(['.sc-13lrnwl-4',  '.sc-pr5ovy-5',  'h1']);
    // Prix HT : sc-13lrnwl-23 (WiFi/Switch/GW/HDDs) | sc-13lrnwl-19 (câbles listing)
    //           sc-pr5ovy-14 (fiche produit legacy) | sc-y2swsu-4 (fiche produit actuelle)
    const domPrice = first(['.sc-13lrnwl-23', '.sc-13lrnwl-19', '.sc-pr5ovy-14', '.sc-y2swsu-4']);
    // Description : sc-13lrnwl-7 (listing) | sc-14cjdti-2 p (fiche produit détaillée)
    const domDesc  = first(['.sc-13lrnwl-7',  '.sc-bw6p3d-13',  '.sc-14cjdti-2 p']);

    // 3. Sold Out (badge OU bouton désactivé)
    const soldEl  = document.querySelector('.sc-1cuaoi0-0');
    const btnSold = document.querySelector('button[label="Sold Out"]');
    const soldOut = (soldEl && soldEl.innerText.toLowerCase().includes('sold')) || !!btnSold;

    // 4. Boutons variante — deux composants possibles selon la page :
    //
    //   a) .sc-14eyr3g-0  : boutons type/longueur/format (Enterprise/Basic, Indoor/Outdoor,
    //                        8TB/16TB/24TB, Professional/Compact, 0.1m/0.3m/1m…)
    //      → présent sur fiches produit ET cartes listing
    //
    //   b) .sc-1uupl4m-6  : swatches couleur sur les fiches produit détaillées
    //                        (Black/White/Blue, etc.)
    //      → uniquement sur les fiches produit (sur les cartes listing les couleurs
    //        sont déjà dans .sc-14eyr3g-0)
    //
    // La logique multi-dimensionnelle est gérée naturellement :
    // chaque visite d'une variante redécouvre SES propres boutons de sous-variante.

    const varTitles = [];

    // a) Boutons type/longueur dans les conteneurs sc-14eyr3g-0
    for (const ctr of document.querySelectorAll('.sc-14eyr3g-0')) {
        const btns = Array.from(ctr.querySelectorAll('button[title]'));
        if (btns.length > 1) {
            for (const b of btns) {
                const t = b.getAttribute('title') || b.title || '';
                if (t && !varTitles.includes(t)) varTitles.push(t);
            }
        }
    }

    // b) Swatches couleur sur fiches produit (sc-1uupl4m-6 > button[title])
    for (const ctr of document.querySelectorAll('.sc-1uupl4m-6')) {
        const btns = Array.from(ctr.querySelectorAll('button[title]'));
        if (btns.length > 1) {
            for (const b of btns) {
                const t = b.getAttribute('title') || b.title || '';
                if (t && !varTitles.includes(t)) varTitles.push(t);
            }
        }
    }

    return {
        sku:      ldSku  || domSku,
        name:     ldName || domName,
        price:    domPrice,    // avec "€" → _parse_price() côté Python
        priceRaw: ldPrice,     // numérique pur JSON-LD (ex: "264.00")
        desc:     ldDesc || domDesc,
        soldOut,
        varTitles
    };
}"""


# ─────────────────────────────────────────────────────────────────────────────
# Scraper principal
# ─────────────────────────────────────────────────────────────────────────────

def scrape_store(scrape_variants: bool = True, debug: bool = False) -> List[Dict]:
    """
    Phase 1 : lit les cartes produit sur les 17 pages catégorie.
              → collecte nom, SKU, desc, prix, dispo directement depuis la carte.
    Phase 2 : visite UNIQUEMENT les fiches avec variantes (button[label]="Select")
              et les pages /collections/. Chaque variante devient une ligne.
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout  # type: ignore
    except ImportError:
        print("[!] Playwright non installé :")
        print("      pip install playwright && playwright install chromium")
        sys.exit(1)

    products: Dict[str, Dict] = {}          # clé = URL normalisée
    queue:    List[Tuple[str, str, str]] = []  # (url, catégorie, label_variante)
    queued:   set = set()
    processed: set = set()

    def enqueue(url: str, cat: str, var_label: str = "",
               keep_query: bool = False) -> None:
        """Ajoute une URL à la file en évitant les doublons.
        keep_query=True : conserve le ?variant= (variantes câbles/accessoires).
        """
        key = url.rstrip("/") if keep_query else _norm_url(url)
        if key not in queued:
            queued.add(key)
            queue.append((key, cat, var_label))

    def _wait_cards(pg, timeout: int = 15_000) -> bool:
        """Attend que React ait rendu au moins une carte produit (layout 1 ou 2)."""
        try:
            pg.wait_for_selector("a.sc-13lrnwl-14, a.sc-bw6p3d-7", timeout=timeout)
            return True
        except Exception:
            return False

    def _card_to_row(c: dict, cat: str, var_label: str = "") -> dict:
        """Convertit une carte JS en dict produit."""
        sold = c.get("soldOut", False) or c.get("btnLabel") == "Sold Out"
        price = _parse_price(c.get("priceHT") or c.get("priceTTC") or "")
        return {
            "product_code":  c.get("sku",  ""),
            "product_name":  c.get("name", ""),
            "variant_label": var_label,
            "category":      cat,
            "price_eur":     price,
            "description":   c.get("desc", ""),
            "available":     "Non" if sold else "Oui",
            "sold_out":      "Oui" if sold else "Non",
            "product_url":   _norm_url(c["href"]),
        }

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        ctx = browser.new_context(user_agent=BROWSER_UA, locale="en-US")

        # Bloque images / polices / médias — n'affecte pas le JS React
        ctx.route(
            "**/*",
            lambda route: route.abort()
            if route.request.resource_type in ("image", "media", "font")
            else route.continue_()
        )

        page = ctx.new_page()

        # ── Phase 1 : pages catégorie ──────────────────────────────────────
        print(f"[1] Collecte des produits sur {len(STORE_CATEGORIES)} catégories …")

        for cat_name, cat_url in STORE_CATEGORIES:
            print(f"    {cat_name} …", end=" ", flush=True)
            try:
                page.goto(cat_url, wait_until="domcontentloaded", timeout=60_000)

                # Attend le rendu React (remplace time.sleep aveugle)
                if not _wait_cards(page, timeout=15_000):
                    print("(attente React…) ", end="", flush=True)
                    time.sleep(4)

                # Scroll progressif → déclenche le lazy-load
                prev_h = 0
                for _ in range(30):
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                    time.sleep(0.5)
                    h = page.evaluate("document.body.scrollHeight")
                    if h == prev_h:
                        break
                    prev_h = h

                # Bouton "Load More" si présent
                for lm_sel in [
                    "button:has-text('Load More')",
                    "button:has-text('Show More')",
                    "[data-testid='load-more']",
                    "a:has-text('Load More')",
                ]:
                    for _ in range(10):
                        try:
                            if page.is_visible(lm_sel, timeout=700):
                                page.click(lm_sel, timeout=3_000)
                                time.sleep(1.0)
                                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                                _wait_cards(page, timeout=4_000)
                            else:
                                break
                        except Exception:
                            break

                cards = page.evaluate(_JS_CATEGORY_CARDS) or []

                n_need = 0
                for c in cards:
                    nu = _norm_url(c["href"])
                    if nu not in products:
                        products[nu] = _card_to_row(c, cat_name)
                    if c.get("needsVisit") and scrape_variants:
                        enqueue(nu, cat_name)
                        n_need += 1

                print(f"{len(cards)} cartes  ({n_need} à explorer)")

                if debug:
                    slug = cat_name.replace(" ", "_").replace("/", "-").replace(".", "")
                    snap = SCRIPT_DIR / f"debug_cat_{slug}.html"
                    snap.write_text(page.content(), encoding="utf-8")
                    print(f"      → snapshot HTML : {snap.name}")

            except Exception as exc:
                print(f"ERREUR : {exc}")

        print(f"\n    → {len(products)} produits uniques  |  {len(queue)} fiches à explorer\n")

        # ── Phase 2 : fiches avec variantes / pages collection ─────────────
        if queue:
            print(f"[2] Exploration des variantes / collections …")
            idx = 0

            while queue:
                url, cat, var_label = queue.pop(0)
                if url in processed:
                    continue
                processed.add(url)
                idx += 1

                try:
                    # ── Page /collections/ : mini-catégorie ────────────────
                    if "/collections/" in url:
                        page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                        _wait_cards(page, timeout=10_000)

                        sub_cards = page.evaluate(_JS_CATEGORY_CARDS) or []
                        for c in sub_cards:
                            nu = _norm_url(c["href"])
                            if nu not in products:
                                products[nu] = _card_to_row(c, cat)
                            if c.get("needsVisit") and scrape_variants:
                                enqueue(nu, cat)

                    # ── Fiche produit individuelle ─────────────────────────
                    else:
                        page.goto(url, wait_until="domcontentloaded", timeout=30_000)
                        time.sleep(0.4)

                        data = page.evaluate(_JS_PRODUCT_DATA)

                        existing = products.get(url, {})
                        sku  = data.get("sku")  or existing.get("product_code", "")
                        name = data.get("name") or existing.get("product_name", "")
                        desc = data.get("desc") or existing.get("description", "")
                        sold = data.get("soldOut", False)

                        # Prix : JSON-LD (pur numérique) > DOM (avec €)
                        price_raw = (data.get("priceRaw") or "").strip()
                        if price_raw:
                            try:
                                float(price_raw)   # valider numérique
                                price = price_raw
                            except ValueError:
                                price = _parse_price(data.get("price", ""))
                        else:
                            price = _parse_price(data.get("price", ""))
                        price = price or existing.get("price_eur", "")

                        products[url] = {
                            "product_code":  sku,
                            "product_name":  name,
                            "variant_label": var_label,
                            "category":      cat,
                            "price_eur":     price,
                            "description":   desc,
                            "available":     "Non" if sold else "Oui",
                            "sold_out":      "Oui" if sold else "Non",
                            "product_url":   url,
                        }

                        # ── Boutons variante → clic + capture URL ──────────
                        if scrape_variants:
                            var_titles = data.get("varTitles", [])
                            if len(var_titles) > 1:
                                for vtitle in var_titles:
                                    try:
                                        clicked = page.evaluate(
                                            r"""(title) => {
                                                // Cherche dans les conteneurs de variantes :
                                                // .sc-14eyr3g-0 : type/longueur/format
                                                // .sc-1uupl4m-6 : swatches couleur (fiche produit)
                                                const containers = [
                                                    ...document.querySelectorAll('.sc-14eyr3g-0'),
                                                    ...document.querySelectorAll('.sc-1uupl4m-6'),
                                                ];
                                                for (const ctr of containers) {
                                                    for (const btn of ctr.querySelectorAll('button[title]')) {
                                                        if (btn.getAttribute('title') === title
                                                            || btn.title === title) {
                                                            btn.click();
                                                            return true;
                                                        }
                                                    }
                                                }
                                                return false;
                                            }""",
                                            vtitle,
                                        )
                                        if not clicked:
                                            continue

                                        # Laisse React mettre à jour l'URL
                                        # (peut changer le path OU juste le ?variant=)
                                        time.sleep(0.6)

                                        current_full = page.url.rstrip("/")
                                        variant_norm = _norm_url(current_full)

                                        if variant_norm != url:
                                            # Le PATH a changé → produit différent
                                            enqueue(variant_norm, cat, vtitle)
                                        elif current_full != url and "?" in current_full:
                                            # Seul le ?variant= a changé (câbles, accessoires)
                                            # → conserver le query param comme clé unique
                                            enqueue(current_full, cat, vtitle,
                                                    keep_query=True)

                                        # Retour à la fiche principale
                                        page.goto(url, wait_until="domcontentloaded",
                                                  timeout=20_000)
                                        # Attend que les boutons variante soient re-rendus
                                        try:
                                            page.wait_for_selector(
                                                ".sc-14eyr3g-0 button[title]",
                                                timeout=5_000)
                                        except Exception:
                                            time.sleep(0.5)

                                    except Exception:
                                        try:
                                            page.goto(url, wait_until="domcontentloaded",
                                                      timeout=15_000)
                                        except Exception:
                                            pass

                except Exception as exc:
                    print(f"    Erreur [{idx}] {url} : {exc}")

                if idx % 10 == 0:
                    n_ok   = sum(1 for p in products.values() if p.get("available") == "Oui")
                    n_sold = sum(1 for p in products.values() if p.get("sold_out")  == "Oui")
                    print(f"    {idx:3d} traités  |  {n_ok} dispo  {n_sold} sold-out  "
                          f"({len(queue)} restants) …")

        browser.close()

    result = list(products.values())
    # Supprime les lignes sans identifiant (artefacts de navigation)
    result = [r for r in result if r.get("product_code") or r.get("product_name")]

    n_avail   = sum(1 for r in result if r.get("available") == "Oui")
    n_sold    = sum(1 for r in result if r.get("sold_out")  == "Oui")
    n_novprix = sum(1 for r in result if not r.get("price_eur"))
    print(f"\n    ✓ {len(result)} articles  |  disponibles: {n_avail}  "
          f"sold-out: {n_sold}  sans prix: {n_novprix}")
    return result


# ─────────────────────────────────────────────────────────────────────────────
# Export CSV
# ─────────────────────────────────────────────────────────────────────────────

def write_csv(rows: List[Dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        print(f"    ✓ CSV sauvé    : {path}")
    except PermissionError:
        from datetime import datetime
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = path.with_stem(f"{path.stem}_{stamp}")
        with alt_path.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=CSV_FIELDS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)
        print(f"    ⚠  Fichier original verrouillé (ouvert ?)")
        print(f"    ✓ CSV sauvé    : {alt_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Export XLSX  (mise en forme professionnelle)
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(rows: List[Dict], path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("    [!] openpyxl manquant → pip install openpyxl   (Excel ignoré)")
        return

    # ── Palette ───────────────────────────────────────────────────────────
    C_HDR_BG  = "1F3864"
    C_HDR_FG  = "FFFFFF"
    C_ROW_A   = "EBF3FB"
    C_ROW_B   = "FFFFFF"
    C_GRN_BG  = "C6EFCE"; C_GRN_FG = "1E6823"
    C_RED_BG  = "FFCCCC"; C_RED_FG = "9C0006"
    C_GRY_BG  = "F2F2F2"; C_GRY_FG = "888888"
    C_LINK_FG = "0563C1"
    C_CODE_FG = "1F3864"
    C_SUM_BG  = "2E4057"

    def _fill(h): return PatternFill("solid", fgColor=h)
    def _font(bold=False, color="000000", size=10, italic=False, underline=None):
        kw = dict(name="Arial", bold=bold, color=color, size=size, italic=italic)
        if underline: kw["underline"] = underline
        return Font(**kw)
    def _align(h="left", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    _side   = Side(border_style="thin", color="D0D8E4")
    _border = Border(left=_side, right=_side, top=_side, bottom=_side)

    wb = Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # Feuille 1 – Catalogue
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Catalogue"
    ws.sheet_properties.tabColor = "1F3864"
    ws.sheet_view.showGridLines = False

    COL_DEFS = [
        # (header,           clé,             largeur, alignement)
        ("Code Produit",    "product_code",   18, "left"),
        ("Nom Commercial",  "product_name",   36, "left"),
        ("Variante",        "variant_label",  16, "center"),
        ("Catégorie",       "category",       22, "left"),
        ("Prix (EUR)",      "price_eur",      14, "right"),
        ("Description",     "description",    62, "left"),
        ("Disponible",      "available",      13, "center"),
        ("Sold Out",        "sold_out",       13, "center"),
        ("Lien Produit",    "product_url",    22, "center"),
    ]
    N = len(COL_DEFS)

    # En-têtes
    for ci, (label, _, width, _align_h) in enumerate(COL_DEFS, 1):
        c = ws.cell(row=1, column=ci, value=label)
        c.font      = _font(bold=True, color=C_HDR_FG, size=11)
        c.fill      = _fill(C_HDR_BG)
        c.alignment = _align(h="center")
        c.border    = _border
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(N)}1"

    # Données
    for ri, row in enumerate(rows, 2):
        bg = C_ROW_A if ri % 2 == 0 else C_ROW_B
        long_desc = False

        for ci, (_, key, _, h_align) in enumerate(COL_DEFS, 1):
            val  = row.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = _border
            cell.alignment = _align(h=h_align, wrap=(key == "description"))
            cell.fill = _fill(bg)
            cell.font = _font()

            if key == "product_code":
                cell.font = _font(bold=True, color=C_CODE_FG)

            elif key == "variant_label" and val:
                cell.font = _font(italic=True, color="555555")

            elif key == "price_eur":
                if val:
                    try:
                        cell.value         = float(val)
                        cell.number_format = '#,##0.00 "€"'
                        cell.font          = _font(bold=True, color=C_CODE_FG)
                    except ValueError:
                        pass
                else:
                    cell.value = "—"
                    cell.fill  = _fill(C_GRY_BG)
                    cell.font  = _font(color=C_GRY_FG, italic=True)

            elif key == "available":
                if val == "Oui":
                    cell.fill = _fill(C_GRN_BG); cell.font = _font(bold=True, color=C_GRN_FG)
                elif val == "Non":
                    cell.fill = _fill(C_RED_BG); cell.font = _font(bold=True, color=C_RED_FG)
                else:
                    cell.value = "—"; cell.fill = _fill(C_GRY_BG)
                    cell.font  = _font(color=C_GRY_FG, italic=True)

            elif key == "sold_out":
                if val == "Oui":
                    cell.fill = _fill(C_RED_BG); cell.font = _font(bold=True, color=C_RED_FG)
                elif val == "Non":
                    cell.fill = _fill(C_GRN_BG); cell.font = _font(color=C_GRN_FG)
                else:
                    cell.value = "—"; cell.fill = _fill(C_GRY_BG)
                    cell.font  = _font(color=C_GRY_FG, italic=True)

            elif key == "product_url":
                if val:
                    cell.hyperlink = val
                    cell.value     = "→ Voir"
                    cell.font      = _font(color=C_LINK_FG, underline="single")
                else:
                    cell.value = "—"
                    cell.fill  = _fill(C_GRY_BG)
                    cell.font  = _font(color=C_GRY_FG, italic=True)

            elif key == "description" and val and len(val) > 80:
                long_desc = True

        ws.row_dimensions[ri].height = 30 if long_desc else 18

    # ════════════════════════════════════════════════════════════════════════
    # Feuille 2 – Résumé
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Résumé")
    ws2.sheet_properties.tabColor = "0E3460"
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 34
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 26

    def _s(r, c, v, bold=False, fg="000000", bg=None, sz=10, h="left", italic=False, ul=None):
        cell = ws2.cell(row=r, column=c, value=v)
        kw = dict(name="Arial", bold=bold, color=fg, size=sz, italic=italic)
        if ul: kw["underline"] = ul
        cell.font      = Font(**kw)
        cell.alignment = Alignment(horizontal=h, vertical="center")
        if bg: cell.fill = PatternFill("solid", fgColor=bg)
        return cell

    # Titre
    ws2.merge_cells("A1:C1")
    _s(1, 1, "Catalogue Produits Ubiquiti", bold=True, fg="FFFFFF",
       bg="0E3460", sz=14, h="center")
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells("A2:C2")
    _s(2, 1, f"Extrait le {_date.today().strftime('%d/%m/%Y')}  •  Source : eu.store.ui.com",
       fg="888888", sz=9, h="center", italic=True)
    ws2.row_dimensions[2].height = 15

    # Stats globales
    n_total    = len(rows)
    n_avail    = sum(1 for r in rows if r.get("available") == "Oui")
    n_sold     = sum(1 for r in rows if r.get("sold_out")  == "Oui")
    n_variant  = sum(1 for r in rows if r.get("variant_label"))
    n_no_price = sum(1 for r in rows if not r.get("price_eur"))

    ws2.merge_cells("A4:C4")
    _s(4, 1, "Vue d'ensemble", bold=True, fg="FFFFFF", bg=C_SUM_BG, sz=11, h="center")
    ws2.row_dimensions[4].height = 20

    stats = [
        ("Total articles",           n_total,    None,     "000000"),
        ("Disponibles",              n_avail,    C_GRN_BG, C_GRN_FG),
        ("Sold Out",                 n_sold,     C_RED_BG, C_RED_FG),
        ("Variantes incluses",       n_variant,  None,     "555555"),
        ("Sans prix (non listé EU)", n_no_price, C_GRY_BG, C_GRY_FG),
    ]

    for i, (label, val, bg_v, fg_v) in enumerate(stats):
        r = 5 + i
        row_bg = C_ROW_A if i % 2 == 0 else C_ROW_B
        _s(r, 1, label, fg="000000", bg=row_bg, sz=10)
        vc = ws2.cell(row=r, column=2, value=val)
        vc.font      = Font(name="Arial", bold=True, color=fg_v, size=11)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        vc.fill      = PatternFill("solid", fgColor=(bg_v or row_bg))
        if n_total > 0 and val > 0:
            pct = val / n_total
            bar = f"{'█' * int(pct * 20)}{'░' * (20 - int(pct * 20))}  {pct:.0%}"
            bc  = ws2.cell(row=r, column=3, value=bar)
            bc.font      = Font(name="Arial", color=(fg_v or "000000"), size=9)
            bc.fill      = PatternFill("solid", fgColor=(bg_v or row_bg))
            bc.alignment = Alignment(horizontal="left", vertical="center")
        ws2.row_dimensions[r].height = 18

    # Répartition par catégorie
    r_title = 5 + len(stats) + 2
    ws2.merge_cells(f"A{r_title}:C{r_title}")
    _s(r_title, 1, "Répartition par catégorie", bold=True, fg="FFFFFF",
       bg=C_SUM_BG, sz=11, h="center")
    ws2.row_dimensions[r_title].height = 20

    r_hdr = r_title + 1
    for ci, (hdr, ha) in enumerate(
        [("Catégorie", "left"), ("Total", "center"), ("Disponibles", "center")], 1
    ):
        c = ws2.cell(row=r_hdr, column=ci, value=hdr)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor=C_SUM_BG)
        c.alignment = Alignment(horizontal=ha, vertical="center")
    ws2.row_dimensions[r_hdr].height = 18

    cat_stats: Dict[str, Dict] = defaultdict(lambda: {"total": 0, "avail": 0})
    for r in rows:
        cat = r.get("category") or "Autre"
        cat_stats[cat]["total"] += 1
        if r.get("available") == "Oui":
            cat_stats[cat]["avail"] += 1

    for i, (cat, cs) in enumerate(
        sorted(cat_stats.items(), key=lambda x: x[1]["total"], reverse=True)
    ):
        r = r_hdr + 1 + i
        bg = C_ROW_A if i % 2 == 0 else C_ROW_B
        _s(r, 1, cat, fg="000000", bg=bg, sz=10)
        tc = ws2.cell(row=r, column=2, value=cs["total"])
        tc.font      = Font(name="Arial", bold=True, size=10)
        tc.fill      = PatternFill("solid", fgColor=bg)
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ac = ws2.cell(row=r, column=3, value=cs["avail"])
        ac.font      = Font(name="Arial", bold=True, size=10,
                            color=(C_GRN_FG if cs["avail"] else C_GRY_FG))
        ac.fill      = PatternFill("solid", fgColor=(C_GRN_BG if cs["avail"] else bg))
        ac.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 16

    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(path))
        print(f"    ✓ Excel sauvé  : {path}")
    except PermissionError:
        # Le fichier est ouvert dans Excel → sauve sous un nom horodaté
        from datetime import datetime
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        alt_path = path.with_stem(f"{path.stem}_{stamp}")
        wb.save(str(alt_path))
        print(f"    ⚠  Fichier original verrouillé (ouvert dans Excel ?)")
        print(f"    ✓ Excel sauvé  : {alt_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Point d'entrée
# ─────────────────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Scraper catalogue Ubiquiti EU Store → XLSX + CSV",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples :
  python ubiquiti_scraper.py
  python ubiquiti_scraper.py --output D:\\SRP\\produits.xlsx
  python ubiquiti_scraper.py --no-variants     # plus rapide, sans variantes
  python ubiquiti_scraper.py --debug           # sauve snapshots HTML
        """,
    )
    parser.add_argument(
        "--output", "-o", default=DEFAULT_OUTPUT, metavar="FICHIER",
        help=f"Chemin de sortie (défaut : {DEFAULT_OUTPUT})",
    )
    parser.add_argument(
        "--no-variants", action="store_true",
        help="Ne pas suivre les boutons variante (plus rapide)",
    )
    parser.add_argument(
        "--debug", action="store_true",
        help="Sauvegarde des snapshots HTML de chaque catégorie",
    )
    # Alias rétrocompatible (ancienne API)
    parser.add_argument("--browser", "-b", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args()

    print(f"Dossier de sortie : {SCRIPT_DIR}\n")

    rows = scrape_store(
        scrape_variants=not args.no_variants,
        debug=args.debug,
    )

    if not rows:
        print("\n[!] Aucun produit collecté.")
        print("    Vérifiez la connexion Internet et les sélecteurs CSS.")
        print("    Relancez avec --debug pour inspecter l'HTML capturé.")
        sys.exit(1)

    base      = Path(args.output).with_suffix("")
    xlsx_path = base.with_suffix(".xlsx")
    csv_path  = base.with_suffix(".csv")

    print("\n[3] Export …")
    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path)

    n_avail = sum(1 for r in rows if r.get("available") == "Oui")
    n_sold  = sum(1 for r in rows if r.get("sold_out") == "Oui")
    n_nopr  = sum(1 for r in rows if not r.get("price_eur"))

    print(f"\n{'═'*58}")
    print(f"  Fichiers créés dans : {SCRIPT_DIR}")
    print(f"    📊  {xlsx_path.name}")
    print(f"    📄  {csv_path.name}")
    print(f"{'─'*58}")
    print(f"  Total articles      : {len(rows)}")
    print(f"    ├─ Disponibles    : {n_avail}")
    print(f"    ├─ Sold Out       : {n_sold}")
    print(f"    └─ Sans prix      : {n_nopr}")
    print(f"{'═'*58}")


if __name__ == "__main__":
    main()
